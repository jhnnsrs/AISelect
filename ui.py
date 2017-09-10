import random
import sys
from datetime import datetime


from bioimage import *

from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QFileDialog, QVBoxLayout, QPushButton, QWidget, QApplication, QLineEdit, QLabel, QGroupBox, \
    QHBoxLayout, QGridLayout, QMessageBox, QTableWidget, QTableWidgetItem, QSlider, QMainWindow, QDockWidget
import numpy as np
import matplotlib.pyplot as plt

from randcolours import rand_cmap
from settings import Global
from linebuilder import LineBuilder
from roilist import RoiListWindow

import projections, postprocess
from elements import CallbackSelector, Roi, RoiParser, AcquiredData
import matplotlib
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas

import os

import h5py

import pandas as pd


#Initialize QT5
matplotlib.use('Qt5Agg')



class ProjectionsSelector(CallbackSelector):

    def itemsToAdd(self):
        self.newItem("Void", projections.void)
        self.newItem("MaxISP",projections.maxisp)

    def setCallback(self,callback):
        Global.projection = callback
        Global.imageWindow.loadImage()

class PostProcessSelector(CallbackSelector):

    def itemsToAdd(self):
        self.newItem("Void",postprocess.void)
        self.newItem("Sobel", postprocess.sobel)
        self.newItem("Laplace (K=1)",postprocess.laplace)

    def setCallback(self,callback):
        Global.postprocess = callback
        Global.handler.loadImage()


class Handler(object):

    def __init__(self):
        Global.handler = self

    def initRoutine(self):

        self.reset()

        self.loadBioImage(Global.filepath)
        self.loadMeta()
        self.loadImage()

        Global.settingsWindow.fileLoaded()
        Global.handler.loadImage()
        Global.metaWindow.updateUI()

    def setStack(self,level):
        #Global.bioImageFile
        pass

    def reloadChannelSelector(self):
        if Global.settingsWindow.selectB4 != None:
            Global.settingsWindow.selectB4 = B4Selector()


    def reset(self):
        Global.colours = iter((rand_cmap(200)(np.linspace(0, 1, 200))))
        Global.bioImageFile = None
        Global.meta = None
        Global.datalist = []
        Global.roilist = []

        if Global.imageWindow.lineBuilder != None:
            Global.imageWindow.lineBuilder.reset() #ahh soooo dirty
        Global.roiListWindow.listUpdated()
        Global.settingsWindow.deleteStageLayout()

    def loadBioImage(self,filepath):
        Global.bioImageFile = BioImageFile(Global.filepath)
        Global.bioImageFile.setSeries(Global.series).run()

    def loadImage(self):
        Global.image = Global.bioImageFile.getSlicedStack(Global.startstack,Global.endstack)
        Global.projected = Global.projection(Global.image)
        Global.postprocessed = Global.postprocess(Global.projected)

        print(Global.postprocessed.shape)

        Global.imageWindow.displayImage(Global.postprocessed)
        # QMessageBox.about(Global.settingsWindow, "Error", "Please select file first")

        Global.handler.reloadChannelSelector()

    def loadMeta(self,filepath = None):
        if filepath is None:
            # TODO: Handle if meta is corrupt
            Global.meta = Global.bioImageFile.meta
        else:
            Global.meta = BioMeta(Global.filepath)

        Global.filename = Global.meta.getFileName
        Global.startstack = 0
        Global.endstack = Global.meta.sizez


    def roiAddedToList(self,roi):
        self.processRoi(roi)

    def processRoi(self,roi):
        data = RoiParser.parseRoi(roi)
        Global.datalist.append(data)
        Global.roiListWindow.listUpdated()

        #if new roi should be automatically shown
        self.showData(-1)

    def deleteROI(self,index):
        self.deletePatches(index)
        Global.datalist.pop(index)
        Global.roilist.pop(index)
        self.listsUpdated()

    def cleanROIs(self):
        for i in range(len(Global.roilist)):
            self.deletePatches(i)
        Global.datalist.clear()
        Global.roilist.clear()
        self.listsUpdated()

    def deletePatches(self,index):

        for i in Global.roilist[index].patcheslink:
            i.remove()

        Global.imageWindow.updateUI()


    def listsUpdated(self):
        Global.roiListWindow.listUpdated()

    def showData(self,dataindex):
        Global.roiWindow.showData(Global.datalist[dataindex])

    def addRoi(self,roi: Roi):
        Global.roilist.append(roi)
        self.roiAddedToList(roi)

    def setMetaWindow(self):
        Global.metaWindow.updateUI()


class B4Selector(CallbackSelector):

    def itemsToAdd(self):
        channelnames = Global.meta.getChannelNames()
        print(channelnames)
        for index,i in enumerate(channelnames):
            self.newItem(i, index)

    def selectItem(self,index):
        self.setCurrentIndex(index)

    def setCallback(self,callback):
        Global.aischannel = callback

class MapSelector(CallbackSelector):

    def itemsToAdd(self):
        channelMaps = Global.availableReadableMaps
        for index, i in enumerate(channelMaps):
            self.newItem(("").join(i) + " to RGB",index)


    def selectItem(self,index):
        self.setCurrentIndex(index)

    def setCallback(self,callback):
        Global.colorMap = Global.availableMaps[callback]
        Global.handler.loadImage()




class SettingsWindow(QWidget):

    def __init__(self,filepath=None,series=None,meta = None,debug=False):
        super(SettingsWindow,self).__init__()
        self.setWindowTitle("AISelect")
        self.filepath = filepath
        Global.settingsWindow = self
        Global.postprocess = postprocess.void
        Global.projection = projections.maxisp
        self.selectB4 = None
        self.stageBox = None
        self.lineBuilder = None

        self.layout = QVBoxLayout(self)

        self.setUI()

    def setUI(self):

        print(next(Global.colours))

        self.changefilebutton = QPushButton('Change File')
        self.flagslabel = QLabel("Flags")
        self.textbox = QLineEdit(self)
        self.donefilebutton = QPushButton('Done')
        self.changefilebutton.clicked.connect(self.changeFile)
        self.donefilebutton.clicked.connect(self.done)
        self.textbox.textChanged.connect(self.changeFlags)


        self.postprocesslabel = QLabel("Post-Processing")
        self.selectPostProcess = PostProcessSelector()

        self.projectionlabel = QLabel("Projection")
        self.selectProjection = ProjectionsSelector()

        self.mappinglabel = QLabel("ColorMapping")
        self.selectMap = MapSelector()


        self.layout.addWidget(self.changefilebutton)

        self.settingBox = QGroupBox("Picture-Settings")
        self.settingLayout  = QGridLayout()
        self.settingLayout.addWidget(self.flagslabel,0,0)
        self.settingLayout.addWidget(self.textbox,0,1)
        self.settingLayout.addWidget(self.projectionlabel,1,0)
        self.settingLayout.addWidget(self.selectProjection,1,1)
        self.settingLayout.addWidget(self.postprocesslabel,2,0)
        self.settingLayout.addWidget(self.selectPostProcess,2,1)
        self.settingLayout.addWidget(self.mappinglabel,3,0)
        self.settingLayout.addWidget(self.selectMap,3,1)
        self.settingBox.setLayout(self.settingLayout)

        self.layout.addWidget(self.settingBox)

        self.layout.addWidget(self.donefilebutton)
        self.setLayout(self.layout)

    def changeFlags(self):
        text = self.textbox.text()
        flags = text.split(";")
        Global.flags = flags

    def addStageLayout(self):
        self.stageBox = QGroupBox("Stage")
        self.stageLayout = QGridLayout()

        self.minimumbox = QLineEdit(self)
        self.minimumbox.setText(str(Global.startstack))
        self.minimumlabel = QLabel("Min")
        self.minimumslider = QSlider(Qt.Horizontal)
        self.minimumslider.setMinimum(0)
        self.minimumslider.setMaximum(Global.meta.sizez-1)
        self.minimumslider.setValue(0)
        self.minimumslider.valueChanged.connect(self.minimumchanged)
        self.minimumbox.textChanged.connect(self.minimumchanged)
        self.stageLayout.addWidget(self.minimumlabel, 0, 0)
        self.stageLayout.addWidget(self.minimumslider,1,0)
        self.stageLayout.addWidget(self.minimumbox, 1, 1)

        self.maximumbox = QLineEdit(self)
        self.maximumbox.setText(str(Global.endstack))
        self.maximumlabel = QLabel("Max")
        self.maximumslider = QSlider(Qt.Horizontal)
        self.maximumslider.setMinimum(0)
        self.maximumslider.setMaximum(Global.meta.sizez-1)
        self.maximumslider.setValue(Global.meta.sizez-1)
        self.maximumslider.valueChanged.connect(self.maximumchanged)
        self.maximumbox.textChanged.connect(self.maximumchanged)
        self.stageLayout.addWidget(self.maximumlabel, 2, 0)
        self.stageLayout.addWidget(self.maximumslider,3,0)
        self.stageLayout.addWidget(self.maximumbox, 3, 1)

        self.snapshotButton = QPushButton("Snapshot Stage")
        self.snapshotButton.clicked.connect(self.snapshotStage)
        self.stageLayout.addWidget(self.snapshotButton,4,0)

        self.snapshotAISButton = QPushButton("Snapshot AIS")
        self.snapshotAISButton.clicked.connect(self.snapshotAIS)
        self.stageLayout.addWidget(self.snapshotAISButton, 4, 1)

        self.stageBox.setLayout(self.stageLayout)
        self.layout.addWidget(self.stageBox)

        self.setLayout(self.layout)

    def deleteStageLayout(self):
        if self.stageBox != None:
            import sip
            self.layout.removeWidget(self.stageBox)
            sip.delete(self.stageBox)
            self.setLayout(self.layout)
            self.stageBox = None

    def minimumchanged(self, value):
        try:
            value = int(value)
        except ValueError:
            value = 0

        if value != Global.startstack and 0 < value < Global.meta.sizez:
            Global.startstack = value

            if Global.startstack >= Global.endstack:
                Global.startstack = Global.endstack - 1

            self.minimumslider.setValue(Global.startstack)
            self.minimumbox.setText(str(Global.startstack))
            Global.handler.loadImage()

    def maximumchanged(self, value):
        try:
            value = int(value)
        except ValueError:
            value = Global.meta.sizez


        if value != Global.endstack and 0 < value < Global.meta.sizez:
            Global.endstack = value

            if Global.startstack >= Global.endstack:
                Global.endstack = Global.startstack + 1

            self.maximumslider.setValue(Global.endstack)
            self.maximumbox.setText(str(Global.endstack))
            Global.handler.loadImage()


    def snapshotStage(self, *args):

        path, file = os.path.split(Global.filepath)
        data = Global.meta.getFileName()
        data = (data[:115] + '-') if len(data) > 115 else data  # TODO: Create clean Filename according to convention
        dirname = path + "/" + "STAGES " + data
        print(dirname)

        if not os.path.exists(dirname):
            os.mkdir(dirname)

        face_file_name = "STAGE_" + data +  "_stage-{0}-{1}".format(Global.startstack,Global.endstack) + ".jpg"
        path = os.path.join(dirname, face_file_name)
        plt.imsave(path, Global.projected,  dpi=600)
        print("Snapshot saved")


    def snapshotAIS(self, *args):
        path, file = os.path.split(Global.filepath)
        data = Global.meta.getFileName()
        data = (data[:115] + '-') if len(data) > 115 else data  # TODO: Create clean Filename according to convention
        dirname = path + "/" + "STAGES " + data
        print(dirname)

        if not os.path.exists(dirname):
            os.mkdir(dirname)

        face_file_name = "AIS-STAGE_" + data + "_stage-{0}-{1}".format(Global.startstack, Global.endstack) + ".jpg"
        path = os.path.join(dirname, face_file_name)

        Global.imageWindow.fig.savefig(path, dpi=600)
        print("Snapshot saved")





    def fileLoaded(self):
        # Now newest Part of Layout should be Loaded
        if self.selectB4 == None:
            self.selectB4 = B4Selector()
            self.b4Label = QLabel("AIS-Channel")

            self.aisBox = QGroupBox("AIS-Settings")
            self.aissettingLayout = QGridLayout()

            self.aissettingLayout.addWidget(self.b4Label,0,0)
            self.aissettingLayout.addWidget(self.selectB4,0,1)

            self.thresholdLabel = QLabel("Threshold")
            self.minimumslider = QSlider(Qt.Horizontal)
            self.minimumslider.setMinimum(0)
            self.minimumslider.setMaximum(100)
            self.minimumslider.setValue(20)
            self.minimumslider.valueChanged.connect(self.thresholdchanged)
            self.aissettingLayout.addWidget(self.thresholdLabel,1,0)
            self.aissettingLayout.addWidget(self.minimumslider, 1, 1)


            self.aisBox.setLayout(self.aissettingLayout)

            self.layout.addWidget(self.aisBox)
            self.setLayout(self.layout)
        if self.stageBox == None:
            self.addStageLayout()

    def thresholdchanged(self,value):
        Global.threshold = float(value/100)
        print(Global.threshold)

    def done(self):

        path, file = os.path.split(Global.filepath)
        data = Global.meta.getFileName()
        data = (data[:115] + '-') if len(data) > 115 else data #TODO: Create clean Filename according to convention
        Global.dirname = path + "/" + "AIS " + data
        print(Global.dirname)

        if not os.path.exists(Global.dirname):
            os.mkdir(Global.dirname)


        for data in Global.datalist:
                face_file_name = "AIS " + "{0:0=3}".format(data.index) + ".jpg"
                path = os.path.join(Global.dirname, face_file_name)
                plt.imsave(path, data.roiimage)

        print("AIS Pictures saved")

        self.saveHDF5()
        self.saveExcel()

        QMessageBox.about(self, "Done", "Files have been successfully writen")

    def saveExcel(self):
        from openpyxl import Workbook
        wb = Workbook()

        # grab the active worksheet
        ws = wb.active

        # Data can be assigned directly to cells
        ws['A1'] = "AIS Index"
        ws['B1'] = "Length in Pixel"
        ws['C1'] = "Physical Length"
        ws['D1'] = "Picture Length"
        ws['E1'] = "Vector Length"
        ws['F1'] = "Loss"

        for data in Global.datalist:
                face_file_name = "AIS " + "{0:0=3}".format(data.index) + ".jpg"
                pixellength = data.aislength
                physicallength = data.aisphysicallength
                picturelength = data.piclength
                vectorlength = data.vectorlength
                ws.append([face_file_name, pixellength, physicallength,picturelength,vectorlength,picturelength/vectorlength])


        # Save the file
        filepath = os.path.join(Global.dirname, "AIS lengths.xlsx")
        wb.save(filepath)


    def saveHDF5(self):

        rantib = Global.metaWindow.table.item(3,1).text()
        gantib = Global.metaWindow.table.item(4,1).text()
        bantib = Global.metaWindow.table.item(5,1).text()
        comments = Global.metaWindow.table.item(0,1).text()
        date = Global.metaWindow.table.item(1,1).text()
        print(comments)
        data = Global.meta.getFileName()
        data = (data[:115] + '-') if len(data) > 115 else data
        filename = data + ".h5"
        timestamp = "T".join( str( datetime.now() ).split() )

        f = h5py.File(os.path.join(Global.dirname,filename), "w")
        # give the HDF5 root some more attributes
        f.attrs['file_name'] = filename
        f.attrs['file_time'] = timestamp
        f.attrs["channels"] = len(Global.meta.getChannelNames())
        f.attrs["channelnames"] = " , ".join(Global.meta.getChannelNames())
        f.attrs["seriesname"] = Global.meta.getSeriesName()
        f.attrs["physicalsizex"] = Global.meta.physicalsizex
        f.attrs["physicalsizey"] = Global.meta.physicalsizey
        f.attrs["physicalsizexunit"] = Global.meta.physicalsizexunit
        f.attrs["physicalsizeyunit"] = Global.meta.physicalsizexunit
        f.attrs["Image-timestamp"] = date
        f.attrs['creator'] = 'AISelect 1.0'
        f.attrs['script_version'] = '1.1b'
        f.attrs['HDF5_Version'] = h5py.version.hdf5_version
        f.attrs['h5py_version'] = h5py.version.version
        f.attrs["R-Antibody"] = rantib
        f.attrs["G-Antibody"] = gantib
        f.attrs["B-Antibody"] = bantib
        f.attrs["Comments"] = comments
        f.create_dataset("BioImage",data=Global.projected)

        roisgroup = f.create_group('ROIs')
        for roi in Global.roilist:
            print("ROI SAVED")
            roigroup = roisgroup.create_group("Roi " + str(roi.index))
            roigroup.create_dataset("Vectors",data=roi.vectors)
            roigroup.create_dataset("Image",data=roi.roiimage)
            roigroup.create_dataset("Input", data=roi.inputvectors)
            roigroup.attrs["width"] = roi.width
            roigroup.attrs["scale"] = roi.scale
            roigroup.attrs["nvectors"] = roi.nvectors

        datagroup = f.create_group("Data")
        for data in Global.datalist:
            print("Data Parsed")
            dataitem = datagroup.create_group("Roi" + str(data.index))
            dataitem.create_dataset("IntensityCurve",data=data.intensitycurves)
            dataitem.create_dataset("Image", data=data.roiimage)
            dataitem.attrs["index"] = data.index
            dataitem.attrs["piclength"] = data.piclength
            pixelgroup = dataitem.create_group("Pixeldata")
            pixelgroup.attrs["AISstart"] = data.aisstart
            pixelgroup.attrs["AISend"] = data.aisend
            dataitem.attrs["B4-Channel"] = data.b4channel
            dataitem.attrs["Threshold"] = data.threshold
            dataitem.attrs["Method"] =  data.parsingmethod
            dataitem.attrs["Flags"] = " , ".join(data.flags)
            dataitem.attrs["Comment"] = data.comment
            dataitem.attrs["StageSTART"] = data.stagestart
            dataitem.attrs["StageEND"] = data.stageend
            dataitem.attrs["Colour"] = data.colour
            physicalgroup = dataitem.create_group("Physical")
            physicalgroup.attrs["AISPhysicalLength"] = data.aisphysicallength

        print("HDF5 Written")
        f.close()

    def changeFile(self):
        #TODO: CLEAN UP ROUTINE
        Global.handler.cleanROIs()
        Global.filepath, _ = QFileDialog.getOpenFileName()

        print(Global.filepath)
        if Global.filepath != "":
            Global.handler.initRoutine()
            print(Global.dirname)

class ImageWindow(QWidget):

    def __init__(self):
        super(ImageWindow,self).__init__()
        self.setWindowTitle("BioImageFile")
        Global.imageWindow = self
        self.layout = QVBoxLayout(self)

        self.lineBuilder = None
        self.fig = plt.figure()
        self.canvas = FigureCanvas(self.fig)
        self.imageCanvas = self.fig.add_subplot(111)

        self.setUI()

    def lineBuilt(self,image):
        #Global.roiWindow.printPictureCurve(image)
        self.canvas.draw()

    def updateUI(self):
        self.setWindowTitle(Global.filepath.split("/")[-1])
        self.canvas.draw()
        self.show()


    def setUI(self):

        self.layout.addWidget(self.canvas)
        self.setLayout(self.layout)



    def instatiateLineBuilder(self):
        if self.lineBuilder == None:
            self.lineBuilder = LineBuilder(self.imageCanvas,self.fig)
            self.lineBuilder.setImageCallback(self.lineBuilt)
            self.lineBuilder.setRoiCallback(Global.handler.addRoi)

    def displayImage(self,image):
            self.figure = self.imageCanvas.imshow(image)
            self.fig.tight_layout()
            self.imageCanvas.axis("off")
            self.figure.axes.get_xaxis().set_visible(False)
            self.figure.axes.get_yaxis().set_visible(False)
            self.imageCanvas.set_xlim([0, image.shape[0]])
            self.imageCanvas.set_ylim([0, image.shape[1]])
            self.updateUI()
            self.instatiateLineBuilder()


class RoiWindow(QWidget):

    def __init__(self):
        super(RoiWindow,self).__init__()
        Global.roiWindow = self
        self.layout = QVBoxLayout(self)
        self.setWindowTitle("Graph and Flourescence")

        self.fig = plt.figure()
        self.canvas = FigureCanvas(self.fig)
        self.dataCanvas = self.fig.add_subplot(212)
        self.pictureCanvas = self.fig.add_subplot(211)

        self.aislength= []


        self.setUI()

    def showData(self,data: AcquiredData):

        self.pictureCanvas.imshow(data.roiimage)

        self.dataCanvas.clear()
        self.dataCanvas.set_xlim([0, data.piclength])


        self.dataCanvas.plot(data.intensitycurves[:, data.b4channel])

        formatedreallength = "{0:.4f}".format(data.aisphysicallength)
        label = formatedreallength + " " + Global.meta.physicalsizexunit #TODO: CHange to AIS centred


        self.dataCanvas.plot([data.aisstart,data.aisend], [data.threshold, data.threshold], 'k-', lw=2)
        self.dataCanvas.text(data.aisstart + data.aislength/2 - 10, data.threshold + 0.1, label)


        self.updateUI()



    def savePicture(self,picture, index):
        if not os.path.exists(Global.dirname):
            os.mkdir(Global.dirname)
        face_file_name = "AIS " + "{0:0=3}".format(index) + ".jpg"
        path = os.path.join(Global.dirname, face_file_name)
        plt.imsave(path, picture)

    def writeList(self):
        df = pd.DataFrame(self.aislength)
        path = os.path.join(Global.dirname, "AIS lengths.csv")
        df.to_csv(path)

    def updateUI(self):
        self.canvas.draw()
        self.show()


    def setUI(self):

        self.layout.addWidget(self.canvas)
        self.setLayout(self.layout)


class MetaWindow(QWidget):

    def __init__(self):
        super(MetaWindow,self).__init__()
        Global.metaWindow = self
        self.layout = QVBoxLayout(self)
        self.setWindowTitle("Metadata")
        self.setUI()

    def setUI(self):
        self.table = QTableWidget()

        # initiate table
        self.table.setWindowTitle("QTableWidget Example @pythonspot.com")
        self.table.setRowCount(6)
        self.table.setColumnCount(2)

        # set label
        self.table.setHorizontalHeaderLabels(("Parameter;Value").split(";"))

        # set data
        self.table.setItem(0, 0, QTableWidgetItem("Voxelsize"))
        self.table.setItem(1, 0, QTableWidgetItem("AcquisitionDate"))
        self.table.setItem(2, 0, QTableWidgetItem("Stacksize"))
        self.table.setItem(3, 0, QTableWidgetItem("R-Antibody"))
        self.table.setItem(4, 0, QTableWidgetItem("G-Antibody"))
        self.table.setItem(5, 0, QTableWidgetItem("B-Antibody"))

        # tooltip text
        self.table.horizontalHeaderItem(0).setToolTip("Size")
        self.table.horizontalHeaderItem(1).setToolTip("Date")

        # show table
        self.layout.addWidget(self.table)
        self.setLayout(self.layout)

    def setMeta(self):
        pass
        #TODO: Implement META Override

    def updateUI(self):
        if Global.meta != None:
            print("META WAS UPDATED")
            self.table.setItem(0, 1, QTableWidgetItem("{0:.4f}".format(Global.meta.physicalsizex) + " X " + "{0:.4f}".format(Global.meta.physicalsizey)))
            self.table.setItem(1, 1, QTableWidgetItem(str(Global.meta.date)))
            self.table.setItem(2, 1, QTableWidgetItem(str(Global.meta.sizez)))
            self.table.setItem(3, 1, QTableWidgetItem("NOT-SET"))
            self.table.setItem(4, 1, QTableWidgetItem("NOT-SET"))
            self.table.setItem(5, 1, QTableWidgetItem("NOT-SET"))
            self.show()
        # DISPLAY META

h_excepthook = sys.excepthook

def my_exception_hook(exctype, value, traceback):
    # Print the error and traceback
    print(exctype, value, traceback)
    # Call the normal Exception hook after
    h_excepthook(exctype, value, traceback)
    sys.exit(1)

# Set the exception hook to our wrapping function
sys.excepthook = my_exception_hook

class Dock(QWidget):

    def __init__(self, parent=None):
        super(Dock, self).__init__(parent)

        self.meta = MetaWindow()
        self.roilist = RoiListWindow()

        self.docklayout = QVBoxLayout()
        self.docklayout.addWidget(self.roilist)
        self.docklayout.addWidget(self.meta)

        self.setLayout(self.docklayout)



class AISelect(QMainWindow):
    def __init__(self, parent=None):
        super(AISelect, self).__init__(parent)
        bar = self.menuBar()
        file = bar.addMenu("File")
        file.addAction("New")

        self.graph = RoiWindow()
        self.dock = Dock()
        self.lala = SettingsWindow()

        self.roidock = QDockWidget("Graph and Flourescence", self)
        self.leftdock = QDockWidget("RoiList and Meta ")

        self.roidock.setWidget(self.graph)
        self.roidock.setFloating(False)

        self.leftdock.setWidget(self.dock)

        self.addDockWidget(Qt.RightDockWidgetArea, self.roidock)
        self.addDockWidget(Qt.LeftDockWidgetArea, self.leftdock)
        self.setCentralWidget(self.lala)

        self.setWindowTitle("AISelect v.2a")





if __name__ == '__main__':

    app = QApplication(sys.argv)

    handler = Handler()

    image = ImageWindow()
    main = AISelect()

    main.show()
    image.show()

    try:
        sys.exit(app.exec_())
    except:
        print("Exiting")